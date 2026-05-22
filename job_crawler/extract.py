"""
数据提取和解析模块（基于 zl 项目）
"""
import json
import re
from html import unescape
from pathlib import Path
from openpyxl import Workbook, load_workbook

# 表字段
HEADERS = [
    "序号",
    "招聘平台",
    "岗位类型一级",
    "岗位类型二级",
    "岗位名称",
    "岗位类型企业/公务员/事业单位/军队文职",
    "公司名称",
    "公司规模",
    "所在省份",
    "城市",
    "详细地址",
    "学历要求",
    "经验要求",
    "薪资范围",
    "福利标签",
    "工作内容",
    "任职要求",
    "岗位链接",
    "发布时间",
    "投递起始时间",
    "投递截止时间",
    "证书要求",
    "备注（技能要求）",
]

# 省份映射（精简版）
CITY_PROVINCE = {
    '北京': '北京市', '天津': '天津市', '上海': '上海市', '重庆': '重庆市',
    '石家庄': '河北省', '太原': '山西省', '呼和浩特': '内蒙古自治区',
    '沈阳': '辽宁省', '长春': '吉林省', '哈尔滨': '黑龙江省',
    '南京': '江苏省', '杭州': '浙江省', '合肥': '安徽省', '福州': '福建省',
    '南昌': '江西省', '济南': '山东省', '郑州': '河南省', '武汉': '湖北省',
    '长沙': '湖南省', '广州': '广东省', '南宁': '广西壮族自治区',
    '海口': '海南省', '成都': '四川省', '贵阳': '贵州省', '昆明': '云南省',
    '拉萨': '西藏自治区', '西安': '陕西省', '兰州': '甘肃省',
    '西宁': '青海省', '银川': '宁夏回族自治区', '乌鲁木齐': '新疆维吾尔自治区',
    '深圳': '广东省', '珠海': '广东省', '汕头': '广东省', '佛山': '广东省',
    '东莞': '广东省', '中山': '广东省', '苏州': '江苏省', '无锡': '江苏省',
    '宁波': '浙江省', '厦门': '福建省', '青岛': '山东省', '大连': '辽宁省',
    '三亚': '海南省', '香港': '香港特别行政区', '澳门': '澳门特别行政区',
    '台北': '台湾省', '全国': '', 'default': ''
}


def clean_header(value):
    """清理表头"""
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value))


def parse_items(json_data):
    """解析 items 列表"""
    if isinstance(json_data, str):
        json_data = json.loads(json_data)

    if isinstance(json_data, list):
        return json_data

    if not isinstance(json_data, dict):
        return []

    if isinstance(json_data.get("list"), list):
        return json_data["list"]

    items = json_data.get("data", {}).get("list", [])
    if isinstance(items, list):
        return items
    return []


def deep_get(data, path, default=""):
    """深度获取字典值"""
    current = data
    for key in path:
        if not isinstance(current, dict):
            return default
        current = current.get(key)
        if current in (None, ""):
            return default
    return current


def first_value(item, *paths):
    """获取第一个非空值"""
    for path in paths:
        value = deep_get(item, path)
        if value not in (None, ""):
            return value
    return ""


def labels_to_text(values):
    """标签列表转文本"""
    if not values:
        return ""

    result = []
    for value in values:
        if isinstance(value, str):
            text = value
        elif isinstance(value, dict):
            text = (
                value.get("name")
                or value.get("value")
                or value.get("tag")
                or value.get("label")
                or value.get("title")
                or ""
            )
        else:
            text = str(value)

        text = text.strip()
        if text and text not in result:
            result.append(text)
    return " / ".join(result)


def keyword_values_to_text(item):
    """提取福利关键词"""
    keywords = item.get("jobKeyword", {}).get("keywords", [])
    values = []
    for keyword in keywords:
        if not isinstance(keyword, dict):
            continue
        value = str(keyword.get("itemValue") or "").strip()
        if value and value not in values:
            values.append(value)
    return " / ".join(values)


def clean_description(text):
    """清理描述文本"""
    if not text:
        return ""

    text = unescape(str(text))
    text = re.sub(r"(?i)<\s*br\s*/?\s*>", "\n", text)
    text = re.sub(r"(?i)</\s*(div|p|li|section|tr|h[1-6])\s*>", "\n", text)
    text = re.sub(r"(?i)<\s*(div|p|li|section|tr|h[1-6])[^>]*>", "\n", text)
    text = re.sub(r"<[^>]+>", "", text)
    text = text.replace("\r", "\n").replace("\u3000", " ").replace("\xa0", " ")

    lines = []
    for line in text.splitlines():
        line = re.sub(r"[ \t]+", " ", line).strip()
        if not line:
            continue
        if re.match(r"^职位福利[:：]", line):
            continue
        if line not in lines[-2:]:
            lines.append(line)

    return "\n".join(lines)


def normalize_section_title(text):
    """规范化段落标题"""
    title_patterns = [
        (r"^(工作职责|岗位职责|职位描述|工作内容)[:：】\]\s]*", r"\1\n"),
        (r"^(任职要求|任职资格|职位要求|岗位要求|应聘要求|资格要求|任职条件)[:：】\]\s]*", r"\1\n"),
    ]
    for pattern, replacement in title_patterns:
        text = re.sub(pattern, replacement, text)
    return text.strip()


def trim_trailing_noise(text):
    """去除尾部噪音"""
    if not text:
        return ""

    stop_pattern = re.compile(
        r"^(.*福利待遇|公司地址|工作地址|工作地点|工作时间|上班时间|联系电话|联系方式|邮箱|简历投递)[:：]?"
    )
    lines = []
    for line in text.splitlines():
        if stop_pattern.match(line.strip()):
            break
        lines.append(line)
    return "\n".join(lines).strip()


def split_description(text):
    """拆分工作内容和任职要求"""
    text = clean_description(text)
    if not text:
        return "", ""

    requirement_pattern = re.compile(
        r"(任职要求|任职资格|职位要求|岗位要求|应聘要求|资格要求|任职条件)[:：】\]\s]*"
    )
    match = requirement_pattern.search(text)
    if not match:
        return trim_trailing_noise(normalize_section_title(text)), ""

    work_content = trim_trailing_noise(
        normalize_section_title(text[: match.start()].strip())
    )
    requirements = trim_trailing_noise(
        normalize_section_title(text[match.start() :].strip())
    )
    return work_content, requirements


def infer_province(item, city):
    """推断省份"""
    address = first_value(
        item,
        ("jobDetailData", "position", "workLocation", "workAddress"),
        ("jobDetailData", "position", "workLocation", "address"),
    )
    match = re.match(
        r"^(北京市|上海市|天津市|重庆市|[^省]+省|[^区]+自治区|香港特别行政区|澳门特别行政区)",
        str(address),
    )
    if match:
        return match.group(1)
    return CITY_PROVINCE.get(city, "")


def build_record(index, item, job_type_level_1="", job_type_level_2=""):
    """构建单条记录"""
    position = item.get("jobDetailData", {}).get("position", {})
    description = deep_get(position, ("desc", "description"))
    work_content, requirements = split_description(description)

    city = first_value(
        item,
        ("workCity",),
        ("jobRootOrgInfo", "cityName"),
        ("jobDetailData", "position", "workLocation", "positionWorkCity"),
    )

    skills = labels_to_text(
        deep_get(position, ("desc", "labels"), [])
        or item.get("jobSkillTags")
        or item.get("skillLabel")
        or item.get("showSkillTags")
    )
    welfare = keyword_values_to_text(item)

    return {
        "序号": index,
        "招聘平台": "智联招聘",
        "岗位类型一级": job_type_level_1,
        "岗位类型二级": job_type_level_2,
        "岗位名称": first_value(
            item, ("name",), ("jobDetailData", "position", "base", "positionName")
        ),
        "岗位类型企业/公务员/事业单位/军队文职": "企业",
        "公司名称": first_value(item, ("companyName",)),
        "公司规模": first_value(
            item, ("companySize",), ("jobDetailData", "companyProxy", "companySize")
        ),
        "所在省份": infer_province(item, city),
        "城市": city,
        "详细地址": first_value(
            item,
            ("jobDetailData", "position", "workLocation", "workAddress"),
            ("jobDetailData", "position", "workLocation", "address"),
        ),
        "学历要求": first_value(
            item, ("education",), ("jobDetailData", "position", "base", "education")
        ),
        "经验要求": first_value(
            item,
            ("workingExp",),
            ("jobDetailData", "position", "base", "positionWorkingExp"),
        ),
        "薪资范围": first_value(
            item, ("salary60",), ("jobDetailData", "position", "base", "salary")
        ),
        "福利标签": welfare,
        "工作内容": work_content,
        "任职要求": requirements,
        "岗位链接": first_value(
            item,
            ("positionUrl",),
            ("positionURL",),
            ("jobDetailData", "position", "base", "positionUrl"),
        ),
        "发布时间": first_value(
            item,
            ("publishTime",),
            ("jobDetailData", "position", "date", "positionPublishTime"),
            ("jobDetailData", "position", "date", "positionUpdateTimeText"),
        ),
        "投递起始时间": first_value(
            item, ("jobDetailData", "position", "date", "dateStart")
        ),
        "投递截止时间": first_value(
            item, ("jobDetailData", "position", "date", "dateEnd")
        ),
        "证书要求": "",
        "备注（技能要求）": f"技能标签：{skills}" if skills else "",
    }


def export_records(items, output_file, job_type_level_1="", job_type_level_2=""):
    """导出记录到 Excel"""
    output_path = Path(output_file)
    if output_path.exists():
        wb = load_workbook(output_path)
        ws = wb.active
        start_index = ws.max_row
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "提取结果"
        ws.append(HEADERS)
        start_index = 1

    for index, item in enumerate(items, start_index):
        record = build_record(index, item, job_type_level_1, job_type_level_2)
        ws.append([record.get(clean_header(header), "") for header in HEADERS])

    wb.save(output_path)


def extract_to_xlsx(json_data, output_file, job_type_level_1="", job_type_level_2=""):
    """
    提取数据并保存到 Excel
    
    Args:
        json_data: API 返回的 JSON 数据（dict 或 str）
        output_file: 输出文件路径
        job_type_level_1: 岗位类型一级
        job_type_level_2: 岗位类型二级
    
    Returns:
        {"count": 条数, "output_file": 文件路径}
    """
    items = parse_items(json_data)
    export_records(items, output_file, job_type_level_1, job_type_level_2)
    return {
        "count": len(items),
        "output_file": str(output_file),
    }
