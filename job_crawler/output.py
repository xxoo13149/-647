import os
import json
import time
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd

from .constants import EMPTY_CELL_VALUE, OUTPUT_COLUMNS
from .utils import (
    choose_latest_publish_time,
    clean_multiline_text,
    clean_text,
    extract_labeled_value,
    fill_empty,
    infer_province,
    merge_distinct_text,
    normalize_city_name,
    normalize_publish_time_text,
    sanitize_filename,
    split_job_summary,
)

PAGE_SIZE = 20


def normalize_job_record(item: dict[str, Any]) -> dict[str, str]:
    """标准化岗位记录字段。"""
    legacy_ability = clean_multiline_text(str(item.get("能力描述", "")))
    legacy_salary = ""
    legacy_location = ""
    legacy_experience = ""
    legacy_education = ""
    legacy_summary = ""
    legacy_tags = ""
    if legacy_ability:
        legacy_salary = extract_labeled_value(legacy_ability, "薪资")
        legacy_location = extract_labeled_value(legacy_ability, "地点")
        legacy_experience = extract_labeled_value(legacy_ability, "经验")
        legacy_education = extract_labeled_value(legacy_ability, "学历")
        legacy_tags = extract_labeled_value(legacy_ability, "技能标签")
        legacy_summary = extract_labeled_value(legacy_ability, "岗位摘要")

    legacy_work, legacy_requirement = split_job_summary(legacy_summary)

    record = {column: fill_empty(item.get(column, "")) for column in OUTPUT_COLUMNS}

    if record["招聘平台"] == EMPTY_CELL_VALUE:
        record["招聘平台"] = "智联招聘"
    if record["岗位类型一级"] == EMPTY_CELL_VALUE:
        record["岗位类型一级"] = fill_empty(item.get("岗位类型一级", item.get("岗位类别/大类", "")))
    if record["岗位类型二级"] == EMPTY_CELL_VALUE:
        record["岗位类型二级"] = fill_empty(item.get("岗位类型二级", ""))
    if record["岗位类型企业/公务员/事业单位/军队文职"] == EMPTY_CELL_VALUE:
        record["岗位类型企业/公务员/事业单位/军队文职"] = fill_empty(
            item.get("岗位类型企业/公务员/事业单位/军队文职", "企业")
        )
    if record["公司名称"] == EMPTY_CELL_VALUE:
        record["公司名称"] = fill_empty(item.get("招聘单位名称", ""))
    if record["岗位名称"] == EMPTY_CELL_VALUE:
        record["岗位名称"] = fill_empty(item.get("岗位名称", ""))
    if record["投递起始时间"] == EMPTY_CELL_VALUE:
        record["投递起始时间"] = fill_empty(normalize_publish_time_text(item.get("最新发布时间", "")))
    else:
        record["投递起始时间"] = fill_empty(normalize_publish_time_text(record["投递起始时间"]))
    if record["发布时间"] == EMPTY_CELL_VALUE:
        record["发布时间"] = record["投递起始时间"]
    if record["证书要求"] == EMPTY_CELL_VALUE:
        record["证书要求"] = fill_empty(item.get("证书要求", ""))

    fallback_map = {
        "薪资范围": legacy_salary,
        "详细地址": legacy_location,
        "经验要求": legacy_experience,
        "学历要求": legacy_education,
        "工作内容": legacy_work,
        "任职要求": legacy_requirement,
    }
    for column, fallback in fallback_map.items():
        if record[column] == EMPTY_CELL_VALUE and fallback:
            record[column] = fill_empty(fallback)

    if record["城市"] == EMPTY_CELL_VALUE:
        record["城市"] = fill_empty(normalize_city_name(record["详细地址"]))
    if record["所在省份"] == EMPTY_CELL_VALUE and record["城市"] != EMPTY_CELL_VALUE:
        record["所在省份"] = fill_empty(infer_province(record["城市"]))

    if legacy_tags:
        remark = "" if record["备注"] == EMPTY_CELL_VALUE else record["备注"]
        tag_remark = "技能标签：" + legacy_tags
        if tag_remark not in remark:
            record["备注"] = fill_empty("；".join([x for x in [remark, tag_remark] if x]))

    if record["公司名称"] == EMPTY_CELL_VALUE:
        record["公司名称"] = "未知单位"
    if record["岗位名称"] == EMPTY_CELL_VALUE:
        record["岗位名称"] = "未知岗位"
    if record["岗位类型一级"] == EMPTY_CELL_VALUE:
        record["岗位类型一级"] = "技术"
    if record["岗位类型企业/公务员/事业单位/军队文职"] == EMPTY_CELL_VALUE:
        record["岗位类型企业/公务员/事业单位/军队文职"] = "企业"

    return record


def load_existing_job_records(file_path: Path) -> list[dict[str, str]]:
    """读取已存在的岗位 Excel 记录。"""
    if not file_path.exists():
        return []

    try:
        sheets = pd.read_excel(file_path, dtype=str, engine="openpyxl", sheet_name=None)
    except Exception as exc:
        print(f"警告：读取现有文件失败，将以新数据重建：{file_path}，原因：{exc}")
        return []

    records = []
    for df in sheets.values():
        if df is None:
            continue
        df = df.fillna("")
        if df.empty:
            continue
        if "序号" in df.columns:
            df = df.drop(columns=["序号"])
        for _, row in df.iterrows():
            records.append(normalize_job_record(row.to_dict()))
    return records


def merge_job_records(
    existing_records: list[dict[str, str]],
    new_records: list[dict[str, str]],
) -> tuple[list[dict[str, str]], int, int]:
    """按岗位链接优先去重，缺失链接时回退到公司 + 岗位 + 城市。"""
    merged_map: dict[tuple[str, ...], dict[str, str]] = {}
    ordered_keys: list[tuple[str, ...]] = []

    for record in existing_records:
        normalized = normalize_job_record(record)
        key = build_record_key(normalized)
        if key in merged_map:
            continue
        merged_map[key] = normalized
        ordered_keys.append(key)

    appended_count = 0
    updated_count = 0

    for record in new_records:
        normalized = normalize_job_record(record)
        key = build_record_key(normalized)

        if key not in merged_map:
            merged_map[key] = normalized
            ordered_keys.append(key)
            appended_count += 1
            continue

        current = merged_map[key]
        changed = False

        merged_publish_time = choose_latest_publish_time(
            current["投递起始时间"],
            normalized["投递起始时间"],
        )
        if merged_publish_time and merged_publish_time != current["投递起始时间"]:
            current["投递起始时间"] = fill_empty(merged_publish_time)
            changed = True

        for column in OUTPUT_COLUMNS:
            if column == "投递起始时间":
                continue
            new_value = normalized[column]
            current_value = current[column]
            if new_value == EMPTY_CELL_VALUE:
                continue
            if current_value == EMPTY_CELL_VALUE or new_value != current_value:
                current[column] = new_value
                changed = True

        if changed:
            updated_count += 1

    merged_records = [merged_map[key] for key in ordered_keys]
    return merged_records, appended_count, updated_count


def build_record_key(record: dict[str, str]) -> tuple[str, ...]:
    """构造稳定去重键。"""
    link = clean_text(record.get("岗位链接", ""))
    if link and link != EMPTY_CELL_VALUE:
        return ("link", link)
    return (
        "fallback",
        clean_text(record.get("公司名称", "")),
        clean_text(record.get("岗位名称", "")),
        clean_text(record.get("城市", "")),
    )


def build_fallback_output_path(file_path: Path) -> Path:
    """Build a unique fallback path when the target workbook is locked."""
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = file_path.with_name(f"{file_path.stem}_recovered_{stamp}{file_path.suffix}")
    if not base.exists():
        return base

    for index in range(2, 1000):
        candidate = file_path.with_name(f"{file_path.stem}_recovered_{stamp}_{index}{file_path.suffix}")
        if not candidate.exists():
            return candidate
    return file_path.with_name(f"{file_path.stem}_recovered_{stamp}_{os.getpid()}{file_path.suffix}")


def write_job_records_to_excel(file_path: Path, records: list[dict[str, str]]) -> Path:
    """将岗位记录写入 Excel，并按 20 条一页拆分为多个工作表。"""
    if not records:
        return file_path

    df = pd.DataFrame(records)
    for column in OUTPUT_COLUMNS:
        if column not in df.columns:
            df[column] = ""
    df = df[OUTPUT_COLUMNS].fillna("")

    file_path = Path(file_path)
    temp_path = file_path.with_name(
        f".{file_path.stem}.{os.getpid()}.{int(time.time() * 1000)}.tmp{file_path.suffix}"
    )

    pages = max(1, (len(df) + PAGE_SIZE - 1) // PAGE_SIZE)
    with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
        for page_index in range(pages):
            start = page_index * PAGE_SIZE
            end = start + PAGE_SIZE
            page_df = df.iloc[start:end].copy()
            page_df.insert(0, "序号", range(start + 1, min(end, len(df)) + 1))
            sheet_name = f"第{page_index + 1}页"
            page_df.to_excel(writer, index=False, sheet_name=sheet_name)

    format_output_workbook(temp_path)

    try:
        os.replace(temp_path, file_path)
        return file_path
    except PermissionError as exc:
        fallback_path = build_fallback_output_path(file_path)
        os.replace(temp_path, fallback_path)
        print(
            f"警告：目标 Excel 正在被占用，无法覆盖：{file_path}，原因：{exc}；"
            f"本轮结果已另存为：{fallback_path}"
        )
        return fallback_path
    except OSError as exc:
        fallback_path = build_fallback_output_path(file_path)
        os.replace(temp_path, fallback_path)
        print(
            f"警告：目标 Excel 写入失败：{file_path}，原因：{exc}；"
            f"本轮结果已另存为：{fallback_path}"
        )
        return fallback_path
    finally:
        if temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                pass


def append_jobs_checkpoint(
    jobs: list[dict],
    output_dir: Path,
    keyword: str,
    session_id: str,
    context: dict[str, Any] | None = None,
) -> Path | None:
    """Append normalized jobs to a JSONL checkpoint so long CLI runs can be recovered."""
    if not jobs:
        return None

    checkpoint_dir = Path(output_dir) / "checkpoints"
    checkpoint_dir.mkdir(parents=True, exist_ok=True)
    base_name = sanitize_filename(keyword, fallback="未知关键词")
    checkpoint_path = checkpoint_dir / f"{base_name}_{session_id}.jsonl"
    context = context or {}
    checkpoint_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with checkpoint_path.open("a", encoding="utf-8") as handle:
        for item in jobs:
            record = normalize_job_record(item)
            payload = {
                "_checkpoint_time": checkpoint_time,
                "_keyword": keyword,
                "_platform": context.get("platform", ""),
                "_region": context.get("region", ""),
                "_page": context.get("page", ""),
                "_detail_index": context.get("detail_index", ""),
                **record,
            }
            handle.write(json.dumps(payload, ensure_ascii=False) + "\n")

    return checkpoint_path


def format_output_workbook(file_path: Path) -> None:
    """按岗位信息表模板的字段宽度做基础可读性格式化。"""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    wb = load_workbook(file_path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    width_by_header = {
        "序号": 8,
        "招聘平台": 12,
        "岗位类型一级": 14,
        "岗位类型二级": 14,
        "岗位名称": 28,
        "岗位类型企业/公务员/事业单位/军队文职": 18,
        "公司名称": 26,
        "公司规模": 14,
        "所在省份": 14,
        "城市": 12,
        "详细地址": 32,
        "学历要求": 12,
        "经验要求": 12,
        "薪资范围": 16,
        "福利标签": 24,
        "工作内容": 48,
        "任职要求": 48,
        "岗位链接": 44,
        "发布时间": 18,
        "投递起始时间": 18,
        "投递截止时间": 18,
        "证书要求": 16,
        "备注": 30,
    }

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for index, cell in enumerate(ws[1], start=1):
            letter = get_column_letter(index)
            ws.column_dimensions[letter].width = width_by_header.get(str(cell.value), 16)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(file_path)


def save_jobs_by_keyword(jobs: list[dict], output_dir: Path, keyword: str) -> dict[str, Any]:
    """按用户输入关键词分文件保存，并执行增量去重更新。"""
    if not jobs:
        return {
            "file_count": 0,
            "raw_count": 0,
            "appended_count": 0,
            "updated_count": 0,
            "saved_files": [],
        }

    base_name = sanitize_filename(keyword, fallback="未知关键词")
    file_name = f"{base_name}.xlsx"
    file_path = output_dir / file_name

    normalized_jobs = [normalize_job_record(item) for item in jobs]
    existing_records = load_existing_job_records(file_path)
    merged_records, appended_count, updated_count = merge_job_records(
        existing_records=existing_records,
        new_records=normalized_jobs,
    )

    saved_path = write_job_records_to_excel(file_path, merged_records)
    target_text = str(file_path) if saved_path == file_path else f"{saved_path}（主文件被占用，已另存）"
    print(
        f"关键词《{keyword}》写入完成：本轮采集 {len(normalized_jobs)} 条，"
        f"新增入库 {appended_count} 条，更新 {updated_count} 条，"
        f"原有 {len(existing_records)} 条，当前共 {len(merged_records)} 条 -> {target_text}"
    )

    return {
        "file_count": 1,
        "raw_count": len(jobs),
        "appended_count": appended_count,
        "updated_count": updated_count,
        "saved_files": [str(saved_path)],
    }
